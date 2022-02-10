import json
import random
import re
import shutil
from random import randrange
from datetime import datetime as dt
import datetime

import pytz
import requests
import xlrd
from django.contrib import messages
from django.contrib.auth import authenticate, logout, login
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.humanize.templatetags.humanize import intcomma
from django.core import serializers
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import ProtectedError, Q
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template
from django.urls import reverse_lazy
from django.utils.datastructures import MultiValueDictKeyError
from django.views.generic import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from onless import settings
from onless.api import SUCCESS, FAILED, INVALID_NUMBER, MESSAGE_IS_EMPTY, SMS_NOT_FOUND, SMS_SERVICE_NOT_TURNED, \
    GetStatusSms, SendSmsWithPlayMobile
from onless.settings import *
from quiz.models import *
from sign.models import Material
from user.decorators import *

from .decorators import *
from .forms import *
from user.models import *
from video.views import *
from video.models import *
from sign.models import *
from .mixins import AllowedRolesMixin
from .models import *
from django.db import IntegrityError, transaction
import pandas as pd
import calendar
from calendar import month_name

#
from .serializers import CreatePupilSerializer
from .utilities import parse_excel_pupil
from .utils import render_to_pdf
from docxtpl import DocxTemplate
from rest_framework import generics


def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        username = get_pasport(username)
        password = request.POST['password'].replace(' ', '')
        user = authenticate(username=username, password=password)
        if user is not None:
            # if user.is_active and user.school.is_block == False:
            if user.is_active:
                login(request, user)
                return redirect(reverse_lazy('landing:home'))
            else:
                messages.error(request, 'Siz bloklangansiz !')
                return HttpResponseRedirect('/accounts/login/')
        else:
            messages.error(request, "Login yoki parol noto'g'ri!")
            return HttpResponseRedirect('/accounts/login/')


@login_required
def add_list(request):
    if request.user.role == '2' or request.user.role == '3':
        return render(request, 'inc/add_list.html')
    else:
        return render(request, 'inc/404.html')


@login_required
def settings_list(request):
    return render(request, 'user/settings_list.html')


@login_required
def worker_add(request):
    if request.user.role == '2':
        if request.method == 'POST':
            form = AddUserForm(data=request.POST)
            password = random.randint(1000000, 9999999)
            school = get_object_or_404(School, id=request.user.school.id)
            if form.is_valid():
                name = form.cleaned_data['name']
                name = get_name(name)
                username = get_pasport(request.POST['pasport'])
                worker = request.POST.get('worker')
                if worker == 'instructor':
                    role = 6
                elif worker == 'accountant':
                    role = 5
                else:
                    role = 3

                try:
                    user = User.objects.create_user(
                        username=username,
                        pasport=username,
                        school=school,
                        turbo=password,
                        password=password,
                        name=name,
                        phone=form.cleaned_data['phone'],
                        role=str(role),
                        is_superuser=False,
                    )
                    user.set_password(password)
                    user.username = username
                    user.email = ''
                    user.save()

                    if school.send_sms_add_worker:
                        r = SendSmsWithApi(user=user, is_add_worker=True).get()
                        print(f"response: {r}")
                        if r == SUCCESS:
                            messages.success(request,
                                             f"{user.name} muvaffaqiyatli qo'shildi! Sms muvaffaqiyatli jo'natildi!")
                        elif r == INVALID_NUMBER:
                            user.delete()
                            messages.error(request, f"{user.phone} raqam noto'g'ri kiritilgan!")
                        elif r == MESSAGE_IS_EMPTY:
                            user.delete()
                            messages.error(request, "Xabar matni kiritilmagan!")
                        elif r == SMS_NOT_FOUND:
                            user.delete()
                            messages.error(request, "Sizda sms to'plami mavjud emas!")
                        elif r == SMS_SERVICE_NOT_TURNED:
                            user.delete()
                            messages.error(request, "Sms xizmati faollashtirilmagan!")
                        else:
                            user.delete()
                            messages.error(request, "Xatolik yuz berdi!")
                    else:
                        messages.success(request,
                                         f"{user.name} muvaffaqiyatli qo'shildi! Sms xizmati o'chirilganligi sababli sms yuborilmadi")
                except IntegrityError:
                    messages.error(request, "Bu pasport oldin ro'yhatdan o'tkazilgan !")
            else:
                messages.error(request, "Forma to'liq yoki to'g'ri to'ldirilmagan !")
        else:
            form = AddUserForm(request)
        return render(request, 'user/worker_add.html', )
    else:
        return render(request, 'inc/404.html')


class AddPupil(AllowedRolesMixin, View):
    allowed_roles = [DIRECTOR, TEACHER]

    def dispatch(self, *args, **kwargs):
        try:
            group = Group.objects.filter(id=kwargs.get('id')).last()
            if self.request.user.school == group.school:
                return super(AllowedRolesMixin, self).dispatch(*args, **kwargs)
            return redirect(reverse_lazy('error_403'))
        except Exception as e:
            print(e)
            return redirect(reverse_lazy('error_403'))

    def get(self, request, *args, **kwargs):
        context = {
            'group_id': self.kwargs.get('id')
        }
        return render(request, 'user/pupil/add_pupil.html', context)

    def post(self, request, *args, **kwargs):
        school = get_object_or_404(School, id=request.user.school.id)
        if school.add_pupil_sms_count <= 0:
            messages.error(request,
                           "Sizda smslar mavjud emas! O'quvchi qo'shish uchun kiritish sms xarid qiling! <a href='{0}'>Sms xarid qilish uchun ushbu havolaga bosing</a>".format(
                               reverse('user:sms_settings')))
            return redirect(reverse_lazy('user:add_pupil', kwargs={'id': self.kwargs.get('id')}))
        serializer = CreatePupilSerializer(data=request.POST,
                                           context={'request': request, 'group_id': kwargs.get('id')})
        if serializer.is_valid():
            serializer.save()
            r = SendSmsWithPlayMobile(user=serializer.instance, is_add_pupil=True).get()
            # r = {'status': SUCCESS, 'result': "Test"}
            if r['status'] == SUCCESS:
                messages.success(request, "O'quvchi muvaffaqiyatli qo'shildi! " + r['result'])
            else:
                messages.warning(request, "O'quvchi muvaffaqiyatli qo'shildi! " + r['result'])
        else:
            try:
                message = next(iter(dict(serializer.errors).values()))[0]
                messages.error(request, message)
            except:
                messages.error(request, serializer.errors)
        return self.get(request, *args, **kwargs)


@login_required
def add_group(request):
    if request.user.role == '2':
        teachers = User.objects.filter(
            Q(role=3, school=request.user.school) | Q(role=2, school=request.user.school)).order_by('name')
        context = {
            'teachers': teachers,
        }
        if request.POST:
            form = AddGroupForm(request.POST)
            if form.is_valid():
                number = form.cleaned_data['number']
                category = request.POST['category']
                teacher = form.cleaned_data['teacher']
                form = form.save(commit=False)
                form.number = number
                if request.POST['price']:
                    price = request.POST['price']
                    form.price = price
                form.category = category
                form.teacher = teacher
                form.school = request.user.school
                form.start = request.POST['start']
                form.stop = request.POST['stop']
                form.save()
                messages.success(request, f"{form.category}-{form.number}-{form.year} guruh qo'shildi")
        else:
            form = AddGroupForm()
        return render(request, 'user/group/add_group.html', context)
    elif request.user.role == '3':
        if request.POST:
            form = AddTeacherGroupForm(request.POST)
            if form.is_valid():
                number = form.cleaned_data['number']
                category = request.POST['category']
                teacher = get_object_or_404(User, id=request.user.id)
                price = form.cleaned_data['price']
                form = form.save(commit=False)
                form.number = number
                form.price = price
                form.category = category
                form.teacher = teacher
                form.school = request.user.school
                form.start = request.POST['start']
                form.stop = request.POST['stop']
                form.save()
                messages.success(request, f"{form.category}-{form.number}-{form.year} guruh qo'shildi")
        else:
            form = AddTeacherGroupForm()
        return render(request, 'user/teacher/add_group.html')
    else:
        return render(request, 'inc/404.html')


@login_required
def groups_list(request):
    if request.user.role == '2':
        groups = Group.objects.filter(school=request.user.school, is_active=True).order_by('sort')
        if not groups.exists():
            messages.error(request, "Guruhlar mavjud emas avval guruh ro'yhatdan o'tkazing !")
        context = {
            'groups': groups,
        }
        return render(request, 'user/group/groups_list.html', context)
    elif request.user.role == '3':
        teacher = User.objects.get(id=request.user.id)
        groups = Group.objects.filter(school=request.user.school, teacher=teacher, is_active=True).order_by('sort')
        if not groups.exists():
            messages.error(request, "Guruhlar mavjud emas avval guruh ro'yhatdan o'tkazing !")
        context = {
            'groups': groups,
        }
        return render(request, 'user/group/groups_list.html', context)
    elif request.user.role == '4':
        pupil = get_object_or_404(User, id=request.user.id)
        groups = Group.objects.filter(id=pupil.group.id, school=request.user.school, is_active=True).order_by('sort')
        if not groups.exists():
            messages.error(request, "Guruhlar mavjud emas avval guruh ro'yhatdan o'tkazing !")

        context = {
            'groups': groups,
        }
        return render(request, 'user/group/groups_list.html', context)


    else:
        return render(request, 'inc/404.html')


@login_required
def group_detail(request, id):
    group = get_object_or_404(Group, id=id)
    if request.user.role == '3':
        if group.teacher != request.user:
            return render(request, 'inc/404.html')

    school = get_object_or_404(School, id=group.school.id)
    if (request.user.school == school and (
            request.user.role == '2' or request.user.role == '3' or request.user.role == '5' or request.user.role == '6')):
        pupils_list = User.objects.filter(role=4, school=request.user.school, group=group, is_active=True).order_by(
            'name')
        # pupils = []
        # for pupil in pupils_list:
        #     get_name(pupil.name.encode('UTF-8'))
        #     pupils.append(pupil)
        context = {
            'group': group,
            'pupils': pupils_list,
        }

        download = request.GET.get('download')
        if download:
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response[
                'Content-Disposition'] = f'attachment; filename="{group.category}-{group.number} {group.year}.xlsx"'

            # create workbook
            wb = Workbook()

            sheet = wb.active
            sheet.alignment = Alignment(horizontal="center", vertical="center")
            # stylize header row
            # 'id','title', 'quantity','pub_date'
            trips_ws = wb.get_sheet_by_name("Sheet")
            trips_ws.column_dimensions['A'].width = 3
            trips_ws.column_dimensions['B'].width = 40
            trips_ws.column_dimensions['C'].width = 15
            trips_ws.column_dimensions['D'].width = 15
            trips_ws.column_dimensions['E'].width = 15

            c1 = sheet.cell(row=1, column=1)

            c1.value = "#"
            c1.font = Font(bold=True)

            c2 = sheet.cell(row=1, column=2)
            c2.value = "F.I.O"
            c2.font = Font(bold=True)

            c3 = sheet.cell(row=1, column=3)
            c3.value = "Tel raqam"
            c3.font = Font(bold=True)

            c4 = sheet.cell(row=1, column=4)
            c4.value = "Login"
            c4.font = Font(bold=True)

            c5 = sheet.cell(row=1, column=5)
            c5.value = "Parol"
            c5.font = Font(bold=True)

            # export data to Excel
            rows = pupils_list.values_list('name', 'phone', 'username', 'turbo')

            for row_num, row in enumerate(rows, 1):
                i = 1
                # row is just a tuple
                for col_num, value in enumerate(row):
                    r2 = sheet.cell(row=row_num + 1, column=col_num + 2)
                    r2.value = value
                    # forloop = sheet.cell(column=1, row=row_num + 1)
                    # forloop.value = '#'

                i = + 1
            wb.save(response)

            return response
        return render(request, 'user/group/group_detail.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def group_delete(request, id):
    director = User.objects.filter(school=request.user.school, role=2).first()
    if request.user == director:
        group = get_object_or_404(Group, id=id)
        pupils = User.objects.filter(school=request.user.school, group=group)
        for pupil in pupils:
            pupil.delete()
        group.delete()
    else:
        return render(request, 'inc/404.html')


@login_required
def group_update(request, id):
    group = get_object_or_404(Group, id=id)
    school = get_object_or_404(School, id=group.school.id)
    if (request.user.school == school and (
            request.user.role == '2' or request.user.role == '3' or request.user.role == '5' or request.user.role == '6')):
        form = GroupUpdateForm(instance=group, user=request.user)
        if request.POST:
            form = GroupUpdateForm(request.POST, instance=group, user=request.user)
            if form.is_valid():
                form.save()
                messages.success(request, "Guruh muvaffaqiyatli tahrirlandi !")
            else:
                messages.error(request, "Formani to'ldirishda xatolik !")
        else:
            form = GroupUpdateForm(instance=group, user=request.user)
        context = {
            'form': form,
            'group': group,
        }
        return render(request, 'user/group/group_update.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def profil_edit(request):
    user = get_object_or_404(User, id=request.user.id)
    form = EditUserForm(instance=user)
    if request.POST:
        form = EditUserForm(request.POST or None, request.FILES or None, instance=user)
        if form.is_valid():
            name = form.cleaned_data['name']
            name = get_name(name)
            form = form.save(commit=False)
            form.name = name
            password = user.set_password(request.POST['turbo'])
            form.save()
            user = authenticate(username=request.user.username, password=request.POST['turbo'])
            if user is not None:
                if user.is_active:
                    login(request, user)
                    form = EditUserForm(instance=user)
            messages.success(request, 'Muvaffaqiyatli tahrirlandi !')
        else:
            messages.error(request, "Formani to'ldirishda xatolik !")
    else:
        form = EditUserForm(instance=user)
    context = {
        'form': form
    }
    return render(request, 'user/profil_edit.html', context)


@login_required
def school_edit(request):
    if request.user.role == '2':
        school = get_object_or_404(School, id=request.user.school.id)
        form = EditSchoolForm(instance=request.user.school)
        if request.POST:
            form = EditSchoolForm(request.POST or None, request.FILES or None, instance=request.user.school)
            if form.is_valid():
                form.save()
                messages.success(request, 'Muvaffaqiyatli tahrirlandi !')
            else:
                messages.error(request, "Formani to'ldirishda xatolik !")
        else:
            form = EditSchoolForm(instance=request.user.school)
        context = {
            'form': form
        }
        return render(request, 'user/school/school_edit.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def pupil_edit(request, id):
    user = get_object_or_404(User, id=id)
    school = get_object_or_404(School, id=user.school.id)
    if (request.user.school == school and (
            request.user.role == '2' or request.user.role == '3' or request.user.role == '6' or request.user.role == '5')):
        groups = Group.objects.filter(is_active=True, school=request.user.school)
        regions = Region.objects.all()
        if request.POST:
            try:
                #
                name = get_name(request.POST.get('name'))
                passport = get_pasport(request.POST['passport'])
                group = get_object_or_404(Group, id=request.POST.get('group'))
                phone = request.POST.get('phone').replace('(', '').replace(')', '').replace('-', '').replace(' ', '')
                password = request.POST.get('password')
                #
                user.name = name
                user.pasport = passport
                user.username = passport
                user.group = group
                user.phone = phone
                user.turbo = password
                user.set_password(password)
                #
                if request.POST.get('birthday'):
                    user.birthday = datetime.datetime.strptime(request.POST.get('birthday', None), '%d.%m.%Y')
                if request.POST.get('place_of_birth'):
                    user.place_of_birth = request.POST.get('place_of_birth')
                if request.POST.get('region'):
                    user.region = get_object_or_404(Region, id=request.POST.get('region'))
                if request.POST.get('district'):
                    user.district = get_object_or_404(District, id=request.POST.get('district'))
                if request.POST.get('residence_address'):
                    user.residence_address = request.POST.get('residence_address')
                #
                if request.POST.get('passport_issued_organization'):
                    user.passport_issued_organization = request.POST.get('passport_issued_organization')
                if request.POST.get('passport_issued_time'):
                    user.passport_issued_time = datetime.datetime.strptime(
                        request.POST.get('passport_issued_time', None), '%d.%m.%Y')
                #
                if request.POST.get('medical_series'):
                    user.medical_series = request.POST.get('medical_series')
                if request.POST.get('medical_issued_organization'):
                    user.medical_issued_organization = request.POST.get('medical_issued_organization')
                if request.POST.get('medical_issued_date'):
                    user.medical_issued_date = datetime.datetime.strptime(
                        request.POST.get('medical_issued_date', None), '%d.%m.%Y')
                #
                if request.POST.get('certificate_series'):
                    user.certificate_series = request.POST.get('certificate_series')
                if request.POST.get('certificate_number'):
                    user.certificate_number = request.POST.get('certificate_number')
                user.save()

                if school.send_sms_edit_pupil:
                    r = SendSmsWithApi(user=user, is_edit_pupil=True).get()
                    print(f"response: {r}")
                    if r == SUCCESS:
                        messages.success(request,
                                         f"{user.name} muvaffaqiyatli tahrirlandi! Sms muvaffaqiyatli jo'natildi!")
                    elif r == INVALID_NUMBER:
                        messages.error(request, f"{user.phone} raqam noto'g'ri kiritilgan!")
                    elif r == MESSAGE_IS_EMPTY:
                        messages.error(request, "Xabar matni kiritilmagan!")
                    elif r == SMS_NOT_FOUND:
                        messages.error(request, "Sizda sms to'plami mavjud emas!")
                    elif r == SMS_SERVICE_NOT_TURNED:
                        messages.error(request, "Sms xizmati faollashtirilmagan!")
                    else:
                        messages.error(request, "SMS yuborishda xatolik yuz berdi!")
                else:
                    messages.success(request,
                                     f"{user.name} muvaffaqiyatli tahrirlandi! Sms xizmati o'chirilganligi sababli sms yuborilmadi")
            except:
                messages.error(request, "Formani to'ldirishda xatolik!")

        context = {
            'user': user,
            'groups': groups,
            'regions': regions
        }
        return render(request, 'user/pupil/pupil_edit.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def contact(request):
    if request.POST or None:
        form = AddContactForm(request.POST or None, request.FILES or None)
        if form.is_valid():
            text = form.cleaned_data['text']
            photo = request.POST['photo']
            form = form.save(commit=False)
            form.text = text
            form.user = request.user
            form.photo = photo
            form.save()
            messages.success(request, "Muvaffaqiyatli jo'natildi !")
        else:
            messages.error(request, "Formani to'ldiring !")
    else:
        form = AddContactForm()
    return render(request, 'user/contact.html')


@login_required
def search(request):
    try:
        query = request.GET.get('q', False).lower()

        context = {
            'q': query,
        }

        count = 0

        if query:
            try:
                teachers = User.objects.filter(
                    Q(pasport__icontains=query) | Q(phone__icontains=query) | Q(username__icontains=query) |
                    Q(name__icontains=query)).filter(role=3, school=request.user.school,
                                                     is_active=True).order_by('name')
                context.update(teachers=teachers)

                count += teachers.count()
            except ValueError:
                pass

            try:
                bugalters = User.objects.filter(
                    Q(pasport__icontains=query) | Q(phone__icontains=query) | Q(username__icontains=query) |
                    Q(name__icontains=query)).filter(role=5, school=request.user.school,
                                                     is_active=True).order_by('name')
                context.update(bugalters=bugalters)
                count += bugalters.count()
            except ValueError:
                pass

            try:
                instructors = User.objects.filter(
                    Q(pasport__icontains=query) | Q(phone__icontains=query) | Q(username__icontains=query) |
                    Q(name__icontains=query)).filter(role=6, school=request.user.school,
                                                     is_active=True).order_by('name')
                context.update(instructors=instructors)
                count += instructors.count()
            except ValueError:
                pass

            try:
                pupils = User.objects.filter(
                    Q(pasport__icontains=query) | Q(phone__icontains=query) | Q(username__icontains=query) |
                    Q(name__icontains=query) &
                    Q(group__isnull=False)).filter(role=4,
                                                   school=request.user.school,
                                                   is_active=True).order_by('name')

                context.update(pupils=pupils)
                count += pupils.count()
            except ValueError:
                pass

            try:
                groups = Group.objects.filter(
                    Q(category__icontains=query) |
                    Q(number=int(query)) |
                    Q(year=int(query))
                ).filter(school=request.user.school)
                context.update(groups=groups)

                count += groups.count()
            except ValueError:
                pass

            try:
                videos = Video.objects.filter(
                    title__icontains=query, is_active=True
                ).distinct()

                context.update(videos=videos)
                count += videos.count()
            except ValueError:
                pass

            try:
                materials = Material.objects.filter(
                    title__icontains=query, is_active=True
                ).distinct()

                context.update(materials=materials)
                count += materials.count()
            except ValueError:
                pass
            context.update(count=count)

        return render(request, 'user/search_result.html', context)
    except AttributeError:
        return redirect(reverse_lazy('video:home'))


@login_required
def pupil_delete(request, id):
    pupil = get_object_or_404(User, id=id)
    school = get_object_or_404(School, id=pupil.school.id)
    if (request.user.school == school and (
            request.user.role == '2' or request.user.role == '3' or request.user.role == '5' or request.user.role == '6')):
        pupil.delete()
        return HttpResponse(status=200)
    else:
        return render(request, 'inc/404.html')


@login_required
def workers_list(request):
    if request.user.role == '2':
        workers = User.objects.filter(
            Q(school=request.user.school, role=5, is_active=True) | Q(school=request.user.school, role=3,
                                                                      is_active=True) | Q(school=request.user.school,
                                                                                          role=6,
                                                                                          is_active=True)).order_by(
            'name')
        if not workers.exists():
            messages.error(request, "Xodimlar ro'yhatga olinmagan!")
        context = {
            'workers': workers
        }
        return render(request, 'user/workers_list.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def worker_edit(request, id):
    worker = get_object_or_404(User, id=id)
    school = get_object_or_404(School, id=worker.school.id)
    if (request.user.school == school and request.user.role == '2' and (
            worker.role == '3' or worker.role == '5' or worker.role == '6')):
        form = EditWorkerForm(instance=worker)
        if request.POST:
            pasport = get_pasport(request.POST['pasport'])
            form = EditWorkerForm(request.POST, request.FILES, instance=worker)
            if form.is_valid():
                form = form.save(commit=False)
                form.pasport = pasport
                form.username = pasport
                worker.set_password(request.POST['turbo'])
                form.save()
                messages.success(request, 'Muvaffaqiyatli tahrirlandi !')

                if school.send_sms_edit_worker:
                    r = SendSmsWithApi(user=worker, is_edit_worker=True).get()
                    print(f"response: {r}")
                    if r == SUCCESS:
                        messages.success(request, f"{worker.name}ga sms muvaffaqiyatli jo'natildi!")
                    elif r == INVALID_NUMBER:
                        messages.error(request, f"{worker.phone} raqam noto'g'ri kiritilgan!")
                    elif r == MESSAGE_IS_EMPTY:
                        messages.error(request, "Xabar matni kiritilmagan!")
                    elif r == SMS_NOT_FOUND:
                        messages.error(request, "Sizda sms to'plami mavjud emas!")
                    elif r == SMS_SERVICE_NOT_TURNED:
                        messages.error(request, "Sms xizmati faollashtirilmagan!")
                    else:
                        messages.error(request, "SMS yuborishda xatolik yuz berdi!")
                else:
                    messages.success(request,
                                     "Xodim muvaffaqiyatli tahrirlandi! Sms xizmati o'chirilganligi sababli sms yuborilmadi")

                form = EditWorkerForm(instance=worker)
            else:
                messages.error(request, "Formani to'ldirishda xatolik !")
        else:
            form = EditWorkerForm(instance=worker)
        context = {
            'worker': worker,
            'form': form
        }
        return render(request, 'user/worker_edit.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def worker_delete(request, id):
    worker = get_object_or_404(User, id=id)
    school = get_object_or_404(School, id=worker.school.id)
    if (request.user.school == school and request.user.role == '2'):
        worker.delete()
    else:
        return render(request, 'inc/404.html')


class CreatePupilWithExcel(AllowedRolesMixin, View):
    allowed_roles = [TEACHER, DIRECTOR]

    def dispatch(self, *args, **kwargs):
        try:
            group = Group.objects.filter(id=kwargs.get('id')).last()
            if self.request.user.school == group.school:
                return super(AllowedRolesMixin, self).dispatch(*args, **kwargs)
            return redirect(reverse_lazy('error_403'))
        except Exception as e:
            print(e)
            return redirect(reverse_lazy('error_403'))

    def post(self, request, *args, **kwargs):
        if request.FILES.get('file'):
            file = request.FILES.get('file')
            if not file.name.endswith(("xlsx")):
                messages.error(request, "xlsx kengaytmadagi excel fayl yuklang!")
                return redirect(reverse_lazy('user:group_detail', kwargs={'id': kwargs.get('id')}))

            datas = parse_excel_pupil(file.file)

            if not len(datas):
                messages.error(request, "Xatolik! Excel faylni to'ldirishda xatolikka yo'l qo'yilgan.")

            if len(datas) > request.user.school.add_pupil_sms_count:
                messages.error(request,
                               "Excel fayldagi o'quvchilar soni {0} ta, sizdagi kiritish smslar soni esa {1} ta! Iltimos yetarlicha kiritish sms xarid qilib qayta urinib ko'ring!".format(
                                   len(datas), request.user.school.add_pupil_sms_count))
                return redirect(reverse_lazy('user:group_detail', kwargs={'id': kwargs.get('id')}))
            i = 0
            row = 1
            for data in datas:
                row += 1
                serializer = CreatePupilSerializer(data=data,
                                                   context={'request': request, 'group_id': kwargs.get('id')})
                if serializer.is_valid():
                    serializer.save()
                    r = SendSmsWithPlayMobile(user=serializer.instance, is_add_pupil=True).get()
                    # r = {'status': SUCCESS, 'result': "Test"}
                    i += 1
                else:
                    try:
                        message = next(iter(dict(serializer.errors).values()))[0]
                        messages.error(request, f"Xatolik! Excel fayldagi {row}-qator. " + message)
                    except:
                        messages.error(request, f"Xatolik! Excel fayldagi {row}-qator. " + serializer.errors)
                    return redirect(reverse_lazy('user:group_detail', kwargs={'id': kwargs.get('id')}))
            if i > 1:
                messages.success(request, "{0} ta o'quvchi muvaffaqiyatli qo'shildi!".format(i))
            return redirect(reverse_lazy('user:group_detail', kwargs={'id': kwargs.get('id')}))
        else:
            messages.error(request, "Fayl yuklanmagan!")
            return redirect(reverse_lazy('user:group_detail', kwargs={'id': kwargs.get('id')}))


@login_required
def upload_file(request):
    if request.user.role == '2' or request.user.role == '3':
        school = get_object_or_404(School, id=request.user.school.id)
        if request.POST and request.FILES:
            file = request.FILES['file']
            file = File.objects.create(file=file)
            if settings.DEBUG:
                file_path = os.path.join(f'{BASE_DIR}{os.path.sep}media{os.path.sep}{file.file}')
            else:
                file_path = os.path.join(
                    f'{os.path.sep}home{os.path.sep}users{os.path.sep}b{os.path.sep}bcloudintelekt{os.path.sep}domains{os.path.sep}onless.uz{os.path.sep}media{os.path.sep}{file.file}')
            group = get_object_or_404(Group, id=request.POST['group'])
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
            number_of_rows = sheet.nrows
            number_of_columns = sheet.ncols
            i = 0
            for row in range(1, number_of_rows):

                name = str((sheet.cell(row, 1).value))
                name = get_name(name)

                pasport = (sheet.cell(row, 2).value)
                pasport = get_pasport(pasport)

                if len(str(pasport)) >= 7 and len(str(pasport)) < 9:
                    messages.error(request, f"Excel fayldagi {row + 1}-ustunda pasport noto'g'ri kiritilgan ! ")
                    break

                if not name:
                    messages.error(request, f"Excel fayldagi {row + 1}-ustunda ism kiritilmagan ! ")
                    break

                if school.send_sms_add_pupil:
                    if school.sms_count == 0:
                        messages.error(request,
                                       "Sizda smslar mavjud emas! O'quvchi qo'shish uchun sms xarid qiling!")
                        return redirect(reverse_lazy('user:group_detail', kwargs={'id': group.id}))

                split_phone = str((sheet.cell(row, 3).value)).split('.')

                try:
                    phone = int(split_phone[0])

                    try:
                        try:
                            parol = random.randint(1000000, 9999999)
                            user = User.objects.create_user(
                                username=pasport,
                                pasport=pasport,
                                school=school,
                                turbo=parol,
                                password=parol,
                                name=name,
                                phone=int(phone),
                                role='4',
                                group=group,
                                is_superuser=False,
                            )
                            user.set_password(parol)
                            user.username = pasport
                            user.email = ''
                            user.save()
                            if school.send_sms_add_pupil:
                                r = SendSmsWithApi(user=user, is_add_pupil=True).get()
                                print(f"response: {r}")
                                if r == SUCCESS:
                                    i = i + 1
                                    messages.success(request, f"{i} ta o'quvchiga sms muvaffaqiyatli jo'natildi!")
                                elif r == INVALID_NUMBER:
                                    user.delete()
                                    messages.error(request, f"{user.phone} raqam noto'g'ri kiritilgan!")
                                elif r == MESSAGE_IS_EMPTY:
                                    user.delete()
                                    messages.error(request, "Xabar matni kiritilmagan!")
                                elif r == SMS_NOT_FOUND:
                                    user.delete()
                                    messages.error(request, "Sizda sms to'plami mavjud emas!")
                                elif r == SMS_SERVICE_NOT_TURNED:
                                    user.delete()
                                    messages.error(request, "Sms xizmati faollashtirilmagan!")
                                else:
                                    user.delete()
                                    messages.error(request, "SMS yuborishda xatolik yuz berdi!")
                            else:
                                messages.success(request,
                                                 "O'quvchi(lar) muvaffaqiyatli qo'shildi! Sms xizmati o'chirilganligi sababli login va parol sms tarzida yuborilmadi!")

                        except IntegrityError:
                            messages.error(request, f"{pasport} pasport oldin ro'yhatdan o'tkazilgan !")
                            next = request.META['HTTP_REFERER']
                            return HttpResponseRedirect(next)


                    except UnboundLocalError:
                        messages.error(request, "Formani to'liq to'ldiring !")
                        next = request.META['HTTP_REFERER']
                        return HttpResponseRedirect(next)


                except ValueError:
                    phone = None

                    try:
                        try:
                            parol = random.randint(1000000, 9999999)
                            user = User.objects.create_user(
                                username=pasport,
                                pasport=pasport,
                                school=request.user.school,
                                turbo=parol,
                                password=parol,
                                name=name,
                                phone=phone,
                                role='4',
                                group=group,
                                is_superuser=False,
                            )
                            user.set_password(parol)
                            user.username = pasport
                            user.email = ''

                            if school.sms_count > 0:
                                school.sms_count = school.sms_count - 2
                                school.save()
                                user.save()
                                messages.success(request,
                                                 "O'quvchi(lar) muvaffaqiyatli qo'shildi!")
                            else:
                                messages.error(request,
                                               "Sizda kiritish smslar mavjud emas! O'quvchi qo'shish uchun kiritish smsidan xarid qiling!")
                                return redirect(reverse_lazy('user:group_detail', kwargs={'id': group.id}))

                        except IntegrityError:
                            messages.error(request, f"{pasport} pasport oldin ro'yhatdan o'tkazilgan !")
                            next = request.META['HTTP_REFERER']
                            return HttpResponseRedirect(next)

                    except UnboundLocalError:
                        messages.error(request, "Formani to'liq to'ldiring !")
                        next = request.META['HTTP_REFERER']
                        return HttpResponseRedirect(next)
            print(i, 973)
            next = request.META['HTTP_REFERER']
            return HttpResponseRedirect(next)
        else:
            messages.error(request, "Faylni kiriting !")
            next = request.META['HTTP_REFERER']
            return HttpResponseRedirect(next)
    else:
        return render(request, 'inc/404.html')


@login_required
def bugalter_groups_list(request):
    try:
        if request.user.role == '2' or request.user.role == '5':
            groups = Group.objects.filter(school=request.user.school, is_active=True).order_by('sort')
        elif request.user.role == '3':
            teacher = get_object_or_404(User, id=request.user.id)
            groups = Group.objects.filter(school=request.user.school, teacher=teacher, is_active=True).order_by('sort')
        else:
            return render(request, 'inc/404.html')

        context = {
            'groups': groups
        }

        try:
            if request.GET.get('startdate') and request.GET.get('startdate') != 'None' and request.GET.get(
                    'startdate') != '':
                context.update(startdate=request.GET.get('startdate'))
                startdate = dt.strptime(request.GET['startdate'], "%Y-%m-%d").replace(tzinfo=ASIA_TASHKENT_TIMEZONE)
                if startdate > timezone.now():
                    context.update(startdate=timezone.now().strftime('%Y-%m-%d'))
                    messages.error(request, 'Boshlanish sana bugungi sanadan katta bo\'lishi mumkin emas!')
                    return render(request, 'user/bugalter/groups_list.html', context)
        except MultiValueDictKeyError:
            pass

        try:
            if request.GET.get('stopdate') and request.GET.get('stopdate') != 'None' and request.GET.get(
                    'stopdate') != '':
                context.update(stopdate=request.GET.get('stopdate'))
                stopdate = dt.strptime(request.GET['stopdate'], "%Y-%m-%d").replace(tzinfo=ASIA_TASHKENT_TIMEZONE)
                if stopdate > timezone.now():
                    context.update(stopdate=timezone.now().strftime('%Y-%m-%d'))
                    messages.error(request, 'Tugash sanasi bugungi sanadan katta bo\'lishi mumkin emas!')
                    return render(request, 'user/bugalter/groups_list.html', context)
        except MultiValueDictKeyError:
            pass

        try:
            if startdate and stopdate:
                if startdate > stopdate:
                    context.update(startdate=timezone.now().strftime('%Y-%m-%d'))
                    context.update(stopdate=timezone.now().strftime('%Y-%m-%d'))
                    messages.error(request, 'Boshlanish sanasi tugash sanasidan katta bo\'lishi mumkin emas!')
                    return render(request, 'user/bugalter/groups_list.html', context)
        except:
            pass

        if not groups.exists():
            messages.error(request, "Guruhlar mavjud emas avval guruh ro'yhatdan o'tkazing !")
        return render(request, 'user/bugalter/groups_list.html', context)

    except:
        messages.error(request, "Xatolik yuz berdi! Sahifani yangilab qayta urinib ko'ring!")
        return redirect(reverse_lazy('user:bugalter_groups_list'))


@login_required
def bugalter_group_detail(request, id):
    group = get_object_or_404(Group, id=id)
    school = get_object_or_404(School, id=group.school.id)
    if (request.user.school == school and (
            request.user.role == '2' or request.user.role == '3' or request.user.role == '5' or request.user.role == '6')):

        pupils = User.objects.filter(role=4, school=request.user.school, group=group, is_active=True).order_by('name')
        total_pay = group.price * pupils.count()
        payments = Pay.objects.filter(pupil__in=pupils, is_active=True)

        total_payments = 0
        for pay in payments:
            total_payments += pay.payment
        total_debit = total_pay - total_payments
        context = {
            'group': group,
            'pupils': pupils,
            'total_pay': total_pay,
            'total_payments': total_payments,
            'total_debit': total_debit
        }
        return render(request, 'user/bugalter/group_detail.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def set_pay(request):
    if request.user.role == '5':
        if request.POST:
            pupil = get_object_or_404(User, id=request.POST['pupil'])
            school = get_object_or_404(School, id=pupil.school.id)
            values = Pay.objects.filter(pupil=pupil, is_active=True)
            payment = 0
            try:
                if request.POST['pay'] <= '0':
                    return HttpResponse(False)
                for value in values:
                    payment += value.payment
                payment += int(request.POST['pay'])
                if pupil.group.price >= payment:
                    obj = Pay.objects.create(pupil=pupil, payment=int(request.POST['pay']))
                    if request.POST.get('comment'):
                        obj.comment = request.POST.get('comment')
                        obj.save()
                    pay = {
                        'payment': payment,
                        'debit': pupil.group.price - payment
                    }

                    if school.send_sms_payment:
                        r = SendSmsWithApi(user=pupil, is_payment=True, pay=request.POST['pay'],
                                           payment=payment).get()
                        print(f"response: {r}")
                        if r == SUCCESS:
                            messages.success(request, f"{pupil.name}ga sms muvaffaqiyatli jo'natildi!")
                        elif r == INVALID_NUMBER:
                            messages.error(request, f"{pupil.phone} raqam noto'g'ri kiritilgan!")
                        elif r == MESSAGE_IS_EMPTY:
                            messages.error(request, "Xabar matni kiritilmagan!")
                        elif r == SMS_NOT_FOUND:
                            messages.error(request, "Sizda sms to'plami mavjud emas!")
                        elif r == SMS_SERVICE_NOT_TURNED:
                            messages.error(request, "Sms xizmati faollashtirilmagan!")
                        else:
                            messages.error(request, "SMS yuborishda xatolik yuz berdi!")

                    return JsonResponse({'pay': pay})
                else:
                    return HttpResponse({pupil.group.price})
            except ValueError:
                return HttpResponse(False)
        return HttpResponse(False)
    else:
        return render(request, 'inc/404.html')


@login_required
def remove_pay(request, id):
    pay = Pay.objects.get(id=id)
    school = get_object_or_404(School, id=pay.pupil.school.id)
    if (request.user.school == school and (request.user.role == '2' or request.user.role == '5') and (
            pay.pupil.role == '4')):
        try:
            payment = pay.payment
            if request.POST.get('removed_reason'):
                pay.removed_reason = request.POST.get('removed_reason')

            pay.is_active = False
            pay.removed_date = timezone.now()
            pay.save()

            return HttpResponse(payment)
        except ObjectDoesNotExist:
            return HttpResponse(False)
    else:
        return render(request, 'inc/404.html')


@login_required
def pay_history(request, user_id, group_id):
    pupil = get_object_or_404(User, id=user_id)
    school = get_object_or_404(School, id=pupil.school.id)

    if (request.user.school == school and (
            request.user.role == '2' or request.user.role == '3' or request.user.role == '6' or request.user.role == '5')) or (
            request.user.school == school and request.user == pupil and request.user.role == '4'):
        payments = Pay.objects.filter(pupil=pupil, is_active=True)

        group = get_object_or_404(Group, id=group_id)

        if request.user.role == '2' or request.user.role == '5':
            removed_pays = Pay.objects.filter(pupil=pupil, is_active=False)
        else:
            removed_pays = None
        context = {
            'pupil': pupil,
            'payments': payments,
            'group': group,
            'removed_pays': removed_pays
        }
        return render(request, 'user/bugalter/pay_history.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def history_view_video_all(request):
    if request.user.role == '2':
        groups = Group.objects.filter(school=request.user.school)
        context = {
            'groups': groups,
        }
        if request.POST:
            if request.POST['startdate']:
                startdate = request.POST['startdate']
                # strokadan time formatga o'tkazish
                # filtrlash uchun soat min secund qo'shish
                startdate = dt.strptime(startdate, '%Y-%m-%d')
                startdate = datetime.datetime.combine(startdate, datetime.time.min)
            else:
                startdate = '2020-06-01'

            context.update(startdate=request.POST['startdate'])

            if request.POST['stopdate']:
                stopdate = request.POST['stopdate']
                # 2020-07-03 23:59:59.999999
                # <class 'datetime.datetime'>
                stopdate = dt.strptime(stopdate, '%Y-%m-%d')
                stopdate = datetime.datetime.combine(stopdate, datetime.time.max)
            else:
                stopdate = datetime.date.today()
                stopdate = datetime.datetime.combine(stopdate, datetime.time.max)

            context.update(stopdate=request.POST['stopdate'])

            if request.POST['group'] != 'False':
                group = get_object_or_404(Group, id=request.POST['group'])
                pupils = User.objects.filter(group=group)
                views = ViewComplete.objects.filter(Q(user__school=request.user.school) & Q(user__in=pupils) & Q(
                    time__range=[startdate, stopdate])).order_by('-time')

                context.update(views=views, guruh=group)
            else:
                views = ViewComplete.objects.filter(Q(user__school=request.user.school) & Q(user__role=4) & Q(
                    time__range=[startdate, stopdate])).order_by('-time')

                context.update(views=views)
            return render(request, 'user/view_video_history_all.html', context)
        else:
            views = ViewComplete.objects.filter(
                Q(user__school=request.user.school, ) & Q(user__role=4) & Q(user__is_active=True)).order_by(
                '-time')

            context.update(views=views)
            return render(request, 'user/view_video_history_all.html', context)

    elif request.user.role == '3':
        teacher = get_object_or_404(User, id=request.user.id)
        groups = Group.objects.filter(teacher=teacher).order_by('sort')
        pupils = User.objects.filter(group__in=groups)
        context = {
            'groups': groups
        }
        if request.POST:
            if request.POST['startdate']:
                startdate = request.POST['startdate']
                # strokadan time formatga o'tkazish
                # filtrlash uchun soat min secund qo'shish
                startdate = dt.strptime(startdate, '%Y-%m-%d')
                startdate = datetime.datetime.combine(startdate, datetime.time.min)
            else:
                startdate = '2020-06-01'

            context.update(startdate=request.POST['startdate'])

            if request.POST['stopdate']:
                stopdate = request.POST['stopdate']
                # 2020-07-03 23:59:59.999999
                # <class 'datetime.datetime'>
                stopdate = dt.strptime(stopdate, '%Y-%m-%d')
                stopdate = datetime.datetime.combine(stopdate, datetime.time.max)
            else:
                stopdate = datetime.date.today()
                stopdate = datetime.datetime.combine(stopdate, datetime.time.max)

            context.update(stopdate=request.POST['stopdate'])

            if request.POST['group'] != 'False':
                group = get_object_or_404(Group, id=request.POST['group'])
                pupils = User.objects.filter(group=group)
                views = ViewComplete.objects.filter(Q(user__school=request.user.school) & Q(user__in=pupils) & Q(
                    time__range=[startdate, stopdate])).order_by('-time')

                context.update(views=views)
                return render(request, 'user/view_video_history_all.html', context)
            views = ViewComplete.objects.filter(Q(user__school=request.user.school) & Q(user__in=pupils) & Q(
                time__range=[startdate, stopdate])).order_by('-time')

            context.update(views=views)
            return render(request, 'user/view_video_history_all.html', context)
        else:
            views = ViewComplete.objects.filter(Q(user__school=request.user.school) & Q(user__in=pupils)).order_by(
                '-time')

            context.update(views=views)
            return render(request, 'user/view_video_history_all.html', context)
    else:
        return render(request, 'inc/404.html')


class HistoryViewVideoAll(LoginRequiredMixin, View):
    model = ViewComplete
    template_name = 'user/view_video_history_all2.html'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    @allowed_users(allowed_roles=[DIRECTOR, TEACHER, INSPECTION, INSTRUCTOR, ACCOUNTANT])
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request, *args, **kwargs):
        print(request.GET)
        return self.get_template()

    def post(self, request, *args, **kwargs):
        print(request.POST)
        return self.get_json_data()

    def get_template(self):
        return render(self.request, self.template_name, self.get_context_data())

    def get_context_data(self, **kwargs):
        context = {
            'groups': Group.objects.filter(Q(is_active=True) & Q(school=self.request.user.school))
        }
        return context

    def get_json_data(self):
        start = int(self.request.POST.get('start'))
        finish = int(self.request.POST.get('limit'))

        qs = self.get_queryset()
        data = self.get_values(qs)

        list_data = []
        for index, item in enumerate(data[start:start + finish], start):
            view = get_object_or_404(self.model, id=item['id'])

            item['id'] = index + 1

            item['time'] = self.MyStrConverter(item['time'])
            if item['user__group']:
                group = get_object_or_404(Group, id=item['user__group'])
                item['user__group'] = f"{group.category}-{group.number} {group.year}"

            if item['video']:
                item['video'] = get_object_or_404(Video, id=item['video']).title
            list_data.append(item)

        context = {
            'length': data.count(),
            'objects': list_data
        }

        data = json.dumps(context)
        return HttpResponse(data, content_type='json')

    def get_queryset(self):
        q = self.request.POST.get('q', '').lower()
        order_by = self.request.POST.get('order_by')
        qs = self.model.objects.filter(Q(user__school=self.request.user.school) & Q(user__role=4)).order_by(
            f"-{order_by}")

        # regex
        full_group_pattern = '^[a-z]{1,}-[0-9]{1,}\s[0-9]{4}$'
        group_category_and_number_pattern = '^[a-z]{1,}-[0-9]{1,}$|^[a-z]{1,}-[0-9]{1,}\s$'
        group_category_pattern_unfinished = '^[a-z]{1,}-$'
        group_year_pattern = '^[0-9]{4}$'
        group_number_pattern = '^[0-9]{1,}$'

        date_pattern = '^[0-9]{2}.[0-9]{2}.[0-9]{4}$'
        day_pattern = '^[0-9]{2}$|^[0-9]{2}.$'
        day_and_month_pattern = '^[0-9]{2}.[0-9]{2}$|^[0-9]{2}.[0-9]{2}.$'

        try:
            if re.match(full_group_pattern, q):
                # print('full_group_pattern')
                category = q.split('-')[0]
                number = q.split('-')[1].split(' ')[0]
                year = q.split('-')[1].split(' ')[1]

                qs = qs.filter(
                    Q(user__group__category=category) & Q(user__group__number=number) & Q(user__group__year=year))
            elif re.match(group_category_and_number_pattern, q):
                # print('group_category_and_number_pattern')
                category = q.split('-')[0]
                number = q.split('-')[1]

                qs = qs.filter(Q(user__group__category__icontains=category) & Q(user__group__number=number))
            elif re.match(group_category_pattern_unfinished, q):
                # print('group_category_pattern_unfinished')
                category = q.split('-')[0]
                qs = qs.filter(Q(user__group__category__icontains=category))
            elif re.match(group_year_pattern, q):
                # print('group_year_pattern')
                qs = qs.filter(user__group__year=q)
            elif re.match(group_number_pattern, q):
                # print('group_number_pattern')
                qs = qs.filter(user__group__number=q)
            elif re.match(date_pattern, q):
                # print('date_pattern')
                date = q[0:10]
                # time = q[11:16]
                # print(date)
                full_date = dt.strptime(q, '%d.%m.%Y').replace(tzinfo=ASIA_TASHKENT_TIMEZONE)
                # print(full_date)
                qs = qs.filter(Q(time__date=full_date.date()))
            elif re.match(day_pattern, q):
                # print('day_pattern')
                day = q[0:2]
                qs = qs.filter(Q(time__day=day))
            elif re.match(day_and_month_pattern, q):
                # print('day_and_month_pattern')
                day = q[0:2]
                month = q[3:5]
                qs = qs.filter(Q(time__day=day) & Q(time__month=month))
            else:
                qs = qs.filter(Q(user__name__icontains=q) | Q(video__title__icontains=q) | Q(user__group__category=q))
            return qs
        except:
            return ViewComplete.objects.none()

    def get_values(self, qs):
        qs = qs.values('id', 'user__name', 'user__group',
                       'video', 'time', )
        return qs

    def MyStrConverter(self, obj):
        if isinstance(obj, (datetime.datetime)):
            return obj.strftime("%d.%m.%Y %H:%M").__str__()
        elif isinstance(obj, (datetime.date)):
            return obj.strftime("%d.%m.%Y").__str__()


@login_required
def history_pupil_view_video(request, id):
    pupil = get_object_or_404(User, id=id)
    school = get_object_or_404(School, id=pupil.school.id)
    if (request.user.school == school and (
            request.user.role == '2' or request.user.role == '3' or request.user.role == '6' or request.user.role == '5')) or (
            request.user.school == school and request.user == pupil and request.user.role == '4'):

        videos = Video.objects.filter(is_active=True).order_by('id')

        context = {
            'videos': videos,
            'pupil': pupil
        }

        context.update(pupil=pupil)
        return render(request, 'user/pupil/history_pupil_view_video.html', context)
    else:
        return render(request, 'inc/404.html')


class HistoryPupilViewVideo(LoginRequiredMixin, DetailView):
    model = ViewComplete
    template_name = 'user/pupil/history_pupil_view_video.html'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        print(kwargs)
        self.pupil = kwargs.get('id')

    @allowed_users(allowed_roles=[DIRECTOR, TEACHER, INSPECTION, INSTRUCTOR, ACCOUNTANT, PUPIL])
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request, *args, **kwargs):
        print(request.GET)
        return self.get_template()

    def get_template(self):
        return render(self.request, self.template_name, self.get_context_data())

    def get_context_data(self, **kwargs):
        context = {
            'videos': Video.objects.filter(is_active=True).order_by('id'),
            'pupil': self.pupil
        }
        return context


@login_required
def get_district(request):
    if request.is_ajax():
        districts = District.objects.filter(region=request.GET.get('region'))
        options = ""
        for district in districts:
            options += f"<option value='{district.id}'>{district.title}</option>"
        return HttpResponse(options)
    else:
        return False


@login_required
def getDistrict(request):
    id = request.GET.get('id', '')
    print(id)
    result = list(District.objects.filter(region_id=int(id)).values('id', 'title'))

    return HttpResponse(json.dumps(result), content_type="application/json")


@login_required
def school_groups(request, id):
    school = get_object_or_404(School, id=id)
    groups = Group.objects.filter(school=school, is_active=True).order_by('sort')
    context = {
        'groups': groups
    }
    return render(request, 'user/inspection/school_groups.html', context)


@login_required
def school_group_detail(request, id):
    group = get_object_or_404(Group, id=id)
    pupils = User.objects.filter(group=group, is_active=True).order_by('name')
    context = {
        'pupils': pupils,
        'group': group
    }
    return render(request, 'user/inspection/school_group_detail.html', context)


@login_required
def support(request):
    return render(request, 'user/messages_part.html')


@login_required
def result(request, id):
    pupil = get_object_or_404(User, id=id)
    if pupil.role == '4':
        group = get_object_or_404(Group, id=pupil.group.id)

    bilets = Bilet.objects.all().count()
    context = {
        'bilets': bilets
    }
    last_check_bilet = CheckTestColor.objects.filter(user=pupil).last()
    if last_check_bilet:
        context.update(last_check_bilet=last_check_bilet.bilet)
    director = User.objects.filter(school=pupil.school, role=2).first()
    try:
        if request.user == pupil:
            return render(request, 'user/result.html', context)
        elif request.user == group.teacher:
            context.update(pupil=pupil)
            return render(request, 'user/result.html', context)
        elif request.user == director:
            context.update(pupil=pupil)
            return render(request, 'user/result.html', context)
        else:
            return render(request, 'inc/404.html')
    except UnboundLocalError:
        context.update(pupil=pupil)
        return render(request, 'user/result.html', context)


@login_required
def news_messages(request):
    return render(request, 'user/news_messages.html')


@login_required
def get_learning_type(request):
    pupil = get_object_or_404(User, id=request.GET.get('pupil'))
    if request.GET.get('checkbox') == 'false':
        pupil.is_offline = False
    elif request.GET.get('checkbox') == 'true':
        pupil.is_offline = True
    else:
        return HttpResponse(False)
    pupil.save()
    return HttpResponse(True)


@login_required
def attendance_groups_list(request):
    groups = Group.objects.filter(Q(is_active=True) & Q(school=request.user.school))
    if not groups.exists():
        messages.error(request, 'Guruhlar mavjud emas!')
        return render(request, 'user/attendance/attendance_groups_list.html')

    if request.user.role == '3':
        groups = Group.objects.filter(Q(is_active=True) & Q(school=request.user.school) & Q(teacher=request.user))
        if not groups.exists():
            messages.error(request, 'Siz rahbarlik qilayotgan guruhlar mavjud emas!')
            return render(request, 'user/attendance/attendance_groups_list.html')

        groups = Group.objects.filter(
            Q(school=request.user.school) & Q(is_active=True) & Q(group_user__is_offline=True) & Q(
                teacher=request.user)).distinct()
        if not groups.exists():
            messages.error(request, 'An\'anaviy ta\'limda o\'qiydigan o\'quvchilar mavjud emas!')
        context = {
            'groups': groups
        }
        return render(request, 'user/attendance/attendance_groups_list.html', context)
    elif request.user.role == '2':
        groups = Group.objects.filter(
            Q(school=request.user.school) & Q(is_active=True) & Q(group_user__is_offline=True)).distinct()
        if not groups.exists():
            messages.error(request, 'An\'anaviy ta\'limda o\'qiydigan o\'quvchilar mavjud emas!')
        context = {
            'groups': groups
        }
        return render(request, 'user/attendance/attendance_groups_list.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def attendance_set_by_group(request, id):
    group = get_object_or_404(Group, id=id)
    school = get_object_or_404(School, id=group.school.id)
    if (request.user.school == school and (
            request.user.role == '2' or request.user.role == '3' or request.user.role == '5' or request.user.role == '6')):

        today_min = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
        today_max = datetime.datetime.combine(datetime.date.today(), datetime.time.max)

        subjects = Subject.objects.filter(
            Q(is_active=True) & Q(categories__title=group.category) & Q(
                subject_schedule__date__range=(today_min, today_max)) & Q(
                subject_schedule__group=group)).distinct()

        if not subjects.exists():
            messages.error(request, f'Jadval bo\'yicha bugunga biriktirilgan fanlar mavjud emas!')
            return redirect(reverse_lazy('user:attendance_groups_list'))
        director = User.objects.filter(school=request.user.school, role=2).first()
        if request.user == group.teacher or request.user == director:
            context = {
                'group': group,
                'subjects': subjects
            }
            return render(request, 'user/attendance/attendance_set_by_group.html', context)
        else:
            messages.error(request,
                           f'{group.category}-{group.number} {group.year} guruhi davomatini belgilash guruh rahbari va maktab rahbariga ruxsat berilgan!')
            return redirect(reverse_lazy('user:attendance_groups_list'))
    else:
        return render(request, 'inc/404.html')


@login_required
def attendance_set_by_subject(request, group_id, subject_id):
    group = get_object_or_404(Group, id=group_id)
    school = get_object_or_404(School, id=group.school.id)
    subject = get_object_or_404(Subject, id=subject_id)
    today_min = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
    today_max = datetime.datetime.combine(datetime.date.today(), datetime.time.max)

    schedules = Schedule.objects.filter(Q(date__range=(today_min, today_max)) & Q(is_active=True) & Q(subject=subject))

    if not schedules.exists():
        messages.error(request, f'Jadval bo\'yicha bugunga biriktirilgan fanlar mavjud emas!')
        return redirect(reverse_lazy('user:attendance_groups_list'))
    director = User.objects.filter(school=request.user.school, role=2).first()
    if (request.user == group.teacher or request.user == director) and request.user.school == school:
        pupils = User.objects.filter(
            Q(school=request.user.school) & Q(is_active=True) & Q(is_offline=True) & Q(group=group))

        if not pupils.exists():
            messages.error(request, f'O\'quvchilar mavjud emas!')
        context = {
            'group': group,
            'subject': subject,
            'pupils': pupils
        }
        return render(request, 'user/attendance/attendance_set_by_subject.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def attendance_set_visited(request):
    if request.POST:
        try:
            pupil = User.objects.get(id=request.POST.get('visited_pupil'))
            subject = Subject.objects.get(id=request.POST.get('visited_subject'))
            date = datetime.datetime.strptime(request.POST.get('visited_date'), "%d.%m.%Y")
            school = School.objects.get(id=pupil.school.id)
            message = ''
            if request.POST.get('visited') == 'False':
                visited = False
            else:
                visited = True

            attendance = Attendance.objects.filter(pupil=pupil, teacher=request.user, subject=subject,
                                                   date=date)

            if not attendance.exists():
                if visited == False:
                    a = Attendance.objects.create(pupil=pupil, teacher=request.user, subject=subject,
                                                  is_visited=visited, date=date, created_date=datetime.datetime.now())

                    if school.send_sms_attendance:
                        r = SendSmsWithApi(user=a.pupil, is_attendance=True, subject=a.subject).get()
                        if r == SUCCESS:
                            message = f"{a.pupil.name}ga sms muvaffaqiyatli jo'natildi!"
                        elif r == INVALID_NUMBER:
                            message = f"{a.pupil.phone} raqam noto'g'ri kiritilgan!"
                        elif r == MESSAGE_IS_EMPTY:
                            message = "Xabar matni kiritilmagan!"
                        elif r == SMS_NOT_FOUND:
                            message = "Sizda sms to'plami mavjud emas!"
                        elif r == SMS_SERVICE_NOT_TURNED:
                            message = "Sms xizmati faollashtirilmagan!"
                        else:
                            message = "SMS yuborishda xatolik yuz berdi!"
                    else:
                        message = "Sms xizmati o'chirilganligi sababli sms jo'natilmadi!"
            else:
                for atten in attendance:
                    atten.is_visited = visited
                    atten.updated_date = datetime.datetime.now()
                    atten.save()

                    if atten.is_visited == False:
                        if school.send_sms_attendance:
                            r = SendSmsWithApi(user=atten.pupil, is_attendance=True, subject=atten.subject).get()
                            if r == SUCCESS:
                                message = f"{atten.pupil.name}ga sms muvaffaqiyatli jo'natildi!"
                            elif r == INVALID_NUMBER:
                                message = f"{atten.pupil.phone} raqam noto'g'ri kiritilgan!"
                            elif r == MESSAGE_IS_EMPTY:
                                message = "Xabar matni kiritilmagan!"
                            elif r == SMS_NOT_FOUND:
                                message = "Sizda sms to'plami mavjud emas!"
                            elif r == SMS_SERVICE_NOT_TURNED:
                                message = "Sms xizmati faollashtirilmagan!"
                            else:
                                message = "SMS yuborishda xatolik yuz berdi!"
                        else:
                            message = "Sms xizmati o'chirilganligi sababli sms jo'natilmadi!"
                    else:
                        atten.delete()

            return HttpResponse(message, status=200)
        except:
            return HttpResponse(status=400)


@login_required
def send_sms(request):
    if request.user.role == '2':
        form = SendSmsForm(data=request.POST)
        groups = Group.objects.filter(school=request.user.school, is_active=True).order_by('-id')
        teachers = User.objects.filter(Q(school=request.user.school) and Q(is_active=True) and Q(role=3)).order_by(
            'name')
        instructors = User.objects.filter(Q(school=request.user.school) and Q(is_active=True) and Q(role=6)).order_by(
            'name')
        accountants = User.objects.filter(Q(school=request.user.school) and Q(is_active=True) and Q(role=5)).order_by(
            'name')
        sms_count = request.user.school.sms_count
        if sms_count == None:
            school_sms_count = 0
        else:
            school_sms_count = sms_count

        context = {
            'groups': groups,
            'school_sms_count': school_sms_count,
            'teachers': teachers,
            'instructors': instructors,
            'accountants': accountants,
            'SMS_PRICE': SMS_PRICE
        }
        if request.method == 'POST':
            if form.is_valid():
                if school_sms_count == 0:
                    messages.error(request, "Sizda sms to'plam mavjud emas!")
                    return render(request, 'user/director/send_sms.html', context)
                text = form.cleaned_data['text']
                message = text.replace('\n', '')  # oxiridagi ortiqcha probellar o'chadi

                try:
                    group = get_object_or_404(Group, id=int(request.POST['group']))
                    users = User.objects.filter(Q(group=group) & Q(is_active=True) & Q(role=4))
                except ValueError:
                    if request.POST.get('group') == 'accountants':
                        users = User.objects.filter(Q(school=request.user.school) & Q(is_active=True) & Q(role=5))
                    elif request.POST.get('group') == 'instructors':
                        users = User.objects.filter(Q(school=request.user.school) & Q(is_active=True) & Q(role=6))
                    else:
                        users = User.objects.filter(Q(school=request.user.school) & Q(is_active=True) & Q(role=3))
                if users.count() <= 0:
                    messages.error(request, "Ushbu guruhda o'quvchi mavjud emas!")
                    return render(request, 'user/director/send_sms.html', context)
                """
                SMS xabarnoma soni, 160 belgidan ko'p bo'lgan hollarda ko'p SMS sarflash tizimi
                """
                # if len(text) >= 0 and len(text) <= 160:
                #     new_sms_count = sms_count - users.count()
                # elif len(text) > 160 and len(text) <= 306:
                #     new_sms_count = sms_count - (users.count() * 2)
                # elif len(text) > 306 and len(text) <= 459:
                #     new_sms_count = sms_count - (users.count() * 3)
                # elif len(text) > 459 and len(text) <= 612:
                #     new_sms_count = sms_count - (users.count() * 4)
                # elif len(text) > 612 and len(text) <= 765:
                #     new_sms_count = sms_count - (users.count() * 5)
                # elif len(text) > 765 and len(text) <= 918:
                #     new_sms_count = sms_count - (users.count() * 6)
                # elif len(text) > 918 and len(text) <= 1071:
                #     new_sms_count = sms_count - (users.count() * 7)
                # elif len(text) > 1071 and len(text) <= 1224:
                #     new_sms_count = sms_count - (users.count() * 8)
                # else:
                #     new_sms_count = sms_count - (users.count() * 30)
                # if new_sms_count >= 0:
                #     Sms.objects.create(sms_count=sms_count - new_sms_count, school=request.user.school, text=text)
                i = 0
                for user in users:
                    i = i + 1
                    r = SendSmsWithApi(user=user, message=message).get()
                    print(f"response: {r}")
                    if r == SUCCESS:
                        user.save()
                        messages.success(request, f"{i} kishiga sms muvaffaqiyatli jo'natildi!")
                    elif r == INVALID_NUMBER:
                        user.delete()
                        messages.error(request, f"{user.phone} raqam noto'g'ri kiritilgan!")
                    elif r == MESSAGE_IS_EMPTY:
                        user.delete()
                        messages.error(request, "Xabar matni kiritilmagan!")
                    elif r == SMS_NOT_FOUND:
                        user.delete()
                        messages.error(request, "Sizda sms to'plami mavjud emas!")
                    elif r == SMS_SERVICE_NOT_TURNED:
                        user.delete()
                        messages.error(request, "Sms xizmati faollashtirilmagan!")
                    else:
                        user.delete()
                        messages.error(request, "SMS yuborishda xatolik yuz berdi!")
                    # Sarflangan smslarni  bazaga yozish
                    # this_user = School.objects.get(school_user=request.user)
                    # this_user.sms_count = new_sms_count
                    # this_user.save()

                    messages.success(request,
                                     f"Sizning SMS xabarnomangiz {i} ta o'quvchiga muvaffaqiyatli yetkazildi!")
                    context.update(school_sms_count=user.school.sms_count)

            else:
                messages.error(request, "Xabar matni kiritilmagan yoki guruh tanlanmagan!")
        else:
            form = SendSmsForm(request)
        return render(request, 'user/director/send_sms.html', context)
    else:
        return render(request, 'inc/404.html')


class SendSmsList(generics.ListAPIView):
    pass


@login_required
def referral_list(request, id):
    if request.user.role == '2' or request.user.role == '5' or request.user.role == '3':
        user = get_object_or_404(User, id=id)
        referrals = Referral.objects.filter(pay__pupil=user)
        # print(referrals)
        if not referrals.exists():
            messages.error(request, "Bonuslar mavjud emas!")
        total_referrals = 0
        for referral in referrals:
            total_referrals += referral.amount
        # print(total_referrals)
        pays = Pay.objects.filter(pupil=user, is_active=True)

        learn_payment = user.group.price
        total_payments = 0
        for pay in pays:
            total_payments += pay.payment
        debt = learn_payment - (total_payments + total_referrals)
        context = {
            'referrals': referrals,
            'pays': pays,
            'total_referrals': total_referrals,
            'total_payments': total_payments,
            'pupil': user,
            'learn_payment': learn_payment,
            'debt': debt
        }
        return render(request, 'user/bugalter/referral_detail.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def rating_groups_list(request):
    today_min = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
    today_max = datetime.datetime.combine(datetime.date.today(), datetime.time.max)

    if request.user.role == '3':
        teacher_groups = Group.objects.filter(
            Q(school=request.user.school) & Q(is_active=True) & Q(teacher=request.user))
        if not teacher_groups.exists():
            messages.error(request, 'Siz rahbarlik qilayotgan guruhlar mavjud emas!')
            return render(request, 'user/rating/rating_groups_list.html')

        groups = Group.objects.filter(
            Q(school=request.user.school) & Q(is_active=True) & Q(group_user__pupil_attendance__is_visited=True) & Q(
                group_user__pupil_attendance__created_date__range=(today_min, today_max)) & Q(
                teacher=request.user)).distinct()
        if not groups.exists():
            messages.error(request, 'Davomat belgilangan guruhlar mavjud emas!')
            return render(request, 'user/rating/rating_groups_list.html')

        context = {
            'groups': groups
        }
        return render(request, 'user/rating/rating_groups_list.html', context)
    elif request.user.role == '2':

        groups = Group.objects.filter(
            Q(school=request.user.school) & Q(is_active=True) & Q(group_user__pupil_attendance__is_visited=True) & Q(
                group_user__pupil_attendance__created_date__range=(today_min, today_max))).distinct()

        if not groups.exists():
            messages.error(request, 'Davomat belgilangan guruhlar mavjud emas!')
        context = {
            'groups': groups
        }
        return render(request, 'user/rating/rating_groups_list.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def rating_set_by_group(request, id):
    group = get_object_or_404(Group, id=id)
    school = get_object_or_404(School, id=group.school.id)
    if (request.user.school == school and (
            request.user.role == '2' or request.user.role == '3' or request.user.role == '5' or request.user.role == '6')):
        today_min = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
        today_max = datetime.datetime.combine(datetime.date.today(), datetime.time.max)

        subjects = Subject.objects.filter(
            Q(is_active=True) & Q(
                subject_schedule__group=group) & Q(subject_schedule__date__range=(today_min, today_max)) & Q(
                subject_attendance__created_date__range=(today_min, today_max))).distinct()

        if not subjects.exists():
            messages.error(request, f'Davomat belgilangan fanlar mavjud emas!')
            return redirect(reverse_lazy('user:rating_groups_list'))
        director = User.objects.filter(school=request.user.school, role=2).first()
        if request.user == group.teacher or request.user == director:
            context = {
                'group': group,
                'subjects': subjects
            }
            return render(request, 'user/rating/rating_set_by_group.html', context)
        else:
            messages.error(request,
                           f'{group.category}-{group.number} {group.year} guruhi baholarini qo\'yish faqatgina guruh rahbari va maktab rahbariga ruxsat berilgan!')
            return redirect(reverse_lazy('user:rating_groups_list'))
    else:
        return render(request, 'inc/404.html')


@login_required
def rating_set_by_subject(request, group_id, subject_id):
    group = get_object_or_404(Group, id=group_id)
    subject = get_object_or_404(Subject, id=subject_id)
    today_min = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
    today_max = datetime.datetime.combine(datetime.date.today(), datetime.time.max)

    schedules = Schedule.objects.filter(Q(date__range=(today_min, today_max)) & Q(is_active=True) & Q(subject=subject))

    if not schedules.exists():
        messages.error(request, f'Jadval bo\'yicha bugunga biriktirilgan fanlar mavjud emas!')
        return redirect(reverse_lazy('user:rating_groups_list'))

    director = User.objects.filter(school=request.user.school, role=2).first()
    if request.user == group.teacher or request.user == director:
        pupils = User.objects.filter(
            Q(school=request.user.school) & Q(is_active=True) & Q(is_offline=True) & Q(group=group) & Q(
                pupil_attendance__subject=subject) & Q(pupil_attendance__is_visited=True) & Q(
                pupil_attendance__updated_date__range=(today_min, today_max)))

        if not pupils.exists():
            messages.error(request, f'O\'quvchilar mavjud emas!')
        context = {
            'group': group,
            'subject': subject,
            'pupils': pupils
        }
        return render(request, 'user/rating/rating_set_by_subject.html', context)
    else:
        return render(request, 'inc/404.html')


@login_required
def rating_create(request):
    if request.POST:
        try:
            pupil = get_object_or_404(User, id=request.POST.get('rating_pupil'))
            subject = get_object_or_404(Subject, id=request.POST.get('rating_subject'))
            school = get_object_or_404(School, id=pupil.school.id)
            date = datetime.datetime.strptime(request.POST.get('rating_date'), "%d.%m.%Y")
            score = request.POST.get('rating')
            message = ''
            ratings = Rating.objects.filter(pupil=pupil, subject=subject, date=date)

            if not ratings.exists():
                rating = Rating.objects.create(score=score, teacher=request.user, subject=subject, pupil=pupil,
                                               date=date,
                                               updated_date=datetime.datetime.now())
                if school.send_sms_rating:
                    r = SendSmsWithApi(user=rating.pupil, is_rating=True, score=rating.score,
                                       subject=rating.subject).get()
                    if r == SUCCESS:
                        message = f"{rating.pupil.name}ga sms muvaffaqiyatli jo'natildi!"
                    elif r == INVALID_NUMBER:
                        message = f"{rating.pupil.phone} raqam noto'g'ri kiritilgan!"
                    elif r == MESSAGE_IS_EMPTY:
                        message = "Xabar matni kiritilmagan!"
                    elif r == SMS_NOT_FOUND:
                        message = "Sizda sms to'plami mavjud emas!"
                    elif r == SMS_SERVICE_NOT_TURNED:
                        message = "Sms xizmati faollashtirilmagan!"
                    else:
                        message = "SMS yuborishda xatolik yuz berdi!"
                else:
                    message = "Sms xizmati o'chirilganligi sababli sms jo'natilmadi!"
            else:
                rating = ratings.first()
                rating.score = score
                rating.updated_date = datetime.datetime.now()
                rating.save()

            return HttpResponse(message, status=200)
        except:
            return HttpResponse(status=400)


@login_required
def electronical_journal(request):
    # # filtered_date = datetime.datetime.strptime(f"01.{month_list[0]}", "%d.%m.%Y")
    # # firstdayofmonth = datetime.datetime.combine(filtered_date, datetime.time.min)
    #
    # # o'quv oyining eng oxirgi kuni
    # endmonth = calendar.monthrange(groups[0].start.year, groups[0].start.month)
    # lastdayofmonth = datetime.datetime(groups[0].start.year, groups[0].start.month, endmonth[1])
    #
    # # last_day = calendar.monthrange(groups[0].stop.year, groups[0].stop.month)
    # # if last_day[1] > groups[0].stop.month:
    # #     lastdayofmonth = groups[0].stop
    # # print(endmonth)
    # # print(lastdayofmonth)
    #
    # group_stop_month = groups[0].stop.strftime("%m")
    # group_stop_day = groups[0].stop.strftime("%d")
    # stop_month = lastdayofmonth.strftime("%m")
    # stop_day = lastdayofmonth.strftime("%d")
    #
    # if group_stop_month == stop_month:
    #     if group_stop_day < stop_day:
    #         lastdayofmonth = groups[0].stop
    #
    # # # guruh o'quv kunlarini olish
    # # get_days = [i.strftime("%d.%m.%Y") for i in pd.date_range(start=groups[0].start, end=lastdayofmonth).tolist()]
    # # # get_days = get_days.apply(lambda x: x[x.dayofweek <= 4])
    # # days_list = list(set(get_days))
    # # days_list.sort(key=lambda date: datetime.datetime.strptime(date, "%d.%m.%Y"))
    # # # # guruh o'quv kunlarini olishni oxiri
    #
    # # guruh o'quv kunlarini olish
    # def daterange(date1, date2):
    #     for n in range(int((date2 - date1).days) + 1):
    #         yield date1 + datetime.timedelta(n)
    #
    # days = []
    # weekdays = [6]
    #
    # for dt in daterange(groups[0].start, lastdayofmonth.date()):
    #     if dt.weekday() not in weekdays:  # to print only the weekdates
    #         days.append(dt.strftime("%d.%m.%Y"))
    # # guruh o'quv kunlarini olishni oxiri
    #
    # users = User.objects.filter(Q(school=request.user.school) & Q(group=groups[0]) & Q(is_active=True))
    #
    # context = {
    #     'groups': groups,
    #     'subjects': subjects,
    #     'cols': days,
    #     'rows': users,
    #     'months_and_years': month_list
    # }
    if request.method == 'GET':
        groups = Group.objects.filter(Q(school=request.user.school) & Q(is_active=True)).order_by('id')
        if not groups.exists():
            messages.error(request, 'Guruhlar mavjud emas!')
            return render(request, 'user/electronical_journal.html')
        subjects = Subject.objects.filter(Q(is_active=True)).distinct()
        if not subjects.exists():
            messages.error(request, 'Fanlar mavjud emas!')
            return render(request, 'user/electronical_journal.html')

        # o'quv oylari yili bilan selectga chiqarish uchun
        get_months_and_years_list = [i.strftime("%m.%Y") for i in
                                     pd.date_range(start=groups[0].start, end=groups[0].stop).tolist()]
        month_list = list(set(get_months_and_years_list))
        month_list.sort(key=lambda date: datetime.datetime.strptime(date, "%m.%Y"))
        # o'quv oylari yili bilan selectga chiqarish uchun oxiri

        if request.GET.get('group'):
            group = get_object_or_404(Group, id=request.GET.get('group'))
        else:
            group = groups.first()

        if request.GET.get('subject'):
            subject = get_object_or_404(Subject, id=request.GET.get('subject'))
        else:
            subject = subjects.first()

        if request.GET.get('month'):
            month_and_year = request.GET.get('month')
        else:
            month_and_year = month_list[0]

        users = User.objects.filter(Q(school=request.user.school) & Q(group=group) & Q(is_active=True))

        # o'quv oylari yili bilan selectga chiqarish uchun
        get_months_and_years_list = [i.strftime("%m.%Y") for i in
                                     pd.date_range(start=group.start, end=group.stop).tolist()]
        month_list = list(set(get_months_and_years_list))
        month_list.sort(key=lambda date: datetime.datetime.strptime(date, "%m.%Y"))
        # o'quv oylari yili bilan selectga chiqarish uchun oxiri

        startdate = datetime.datetime.strptime(month_and_year, '%m.%Y')
        get_month = startdate.strftime("%m")
        get_year = startdate.strftime("%Y")
        get_group_month = group.start.strftime("%m")

        # agar o'qishning birinchi oyi bo'lsa
        if get_group_month == get_month:
            first_day = f"{group.start.strftime('%d')}.{get_month}.{get_year}"
            firstdayofmonth = datetime.datetime.strptime(first_day, '%d.%m.%Y')
        else:
            first_day = f"01.{get_month}.{get_year}"
            firstdayofmonth = datetime.datetime.strptime(first_day, '%d.%m.%Y')

        # o'quv oyining eng oxirgi kuni
        endmonth = calendar.monthrange(int(get_year), int(get_month))
        lastdayofmonth = datetime.datetime(int(get_year), int(get_month), endmonth[1]).date()

        group_stop_month = group.stop.strftime("%m")
        group_stop_day = group.stop.strftime("%d")
        stop_month = lastdayofmonth.strftime("%m")
        stop_day = lastdayofmonth.strftime("%d")

        if group_stop_month == stop_month:
            if group_stop_day < stop_day:
                lastdayofmonth = group.stop

        # # guruh o'quv kunlarini olish
        # get_days = [i.strftime("%d.%m.%Y") for i in pd.date_range(start=firstdayofmonth, end=lastdayofmonth).tolist()]
        # days_list = list(set(get_days))
        # days_list.sort(key=lambda date: datetime.datetime.strptime(date, "%d.%m.%Y"))
        # # guruh o'quv kunlarini olishni oxiri

        def daterange(date1, date2):
            for n in range(int((date2 - date1).days) + 1):
                yield date1 + datetime.timedelta(n)

        # guruh o'quv kunlarini olish
        days = []
        weekdays = [6]

        for dt in daterange(firstdayofmonth.date(), lastdayofmonth):
            if dt.weekday() not in weekdays:  # to print only the weekdates
                days.append(dt.strftime("%d.%m.%Y"))
        # guruh o'quv kunlarini olishni oxiri

        context = {
            'rows': users,
            'cols': days,
            'group': group,
            'subject': subject,
            'subjects': subjects,
            'groups': groups,
            'months_and_years': month_list,
            'selected_months_and_years': month_and_year
        }

        return render(request, 'user/electronical_journal.html', context)


@login_required
def get_group_pupils_count(request):
    if request.POST:
        group = get_object_or_404(Group, id=request.POST.get('group'))
        pupils_count = User.objects.filter(
            Q(is_active=True) & Q(role=4) & Q(group=group) & Q(phone__isnull=False)).count()
        return HttpResponse(pupils_count)
    else:
        return HttpResponse(False)


@login_required
def get_workers_count(request):
    if request.POST:
        context = {}
        if request.POST.get('workers') == 'teachers':
            teachers_count = User.objects.filter(Q(is_active=True) & Q(school=request.user.school) & Q(role=3)).count()
            return HttpResponse(teachers_count)
        elif request.POST.get('workers') == 'instructors':
            instructors_count = User.objects.filter(
                Q(is_active=True) & Q(school=request.user.school) & Q(role=6)).count()
            return HttpResponse(instructors_count)
        elif request.POST.get('workers') == 'accountants':
            accountants_count = User.objects.filter(
                Q(is_active=True) & Q(school=request.user.school) & Q(role=5)).count()
            return HttpResponse(accountants_count)
        else:
            return HttpResponse(False)
    else:
        return HttpResponse(False)


# @login_required
# def get_attendance_time(request):
#     if request.POST:
#         pupil = get_object_or_404(User, id=request.POST.get('pupil'))
#         subject = get_object_or_404(Subject, id=request.POST.get('subject'))
#         today_min = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
#         today_max = datetime.datetime.combine(datetime.date.today(), datetime.time.max)
#         try:
#             attendance = Attendance.objects.filter(pupil=pupil, subject=subject,
#                                                    created_date__range=(today_min, today_max))
#             for atten in attendance:
#                 import pytz
#                 new_timezone = pytz.timezone("Asia/Tashkent")
#                 new_timezone_timestamp = atten.updated_date.astimezone(new_timezone)
#                 get_datetime = new_timezone_timestamp.strftime('%d.%m.%Y %H:%M')
#                 return HttpResponse(get_datetime)
#         except:
#             return HttpResponse(False)
#     else:
#         return HttpResponse(False)


@login_required
def get_group_months(request):
    if request.is_ajax():
        if request.POST:
            group = get_object_or_404(Group, id=request.POST.get('group'))
            get_months_and_years_list = [i.strftime("%m.%Y") for i in
                                         pd.date_range(start=group.start, end=group.stop).tolist()]
            month_list = list(set(get_months_and_years_list))
            month_list.sort(key=lambda date: datetime.datetime.strptime(date, "%m.%Y"))
            # for get_list in month_list:
            #     split_month = str(get_list).split(' ')[0]
            # group_months = [calendar.month_name[int(i)] for i in month_list]
            # month_lookup = list(month_name)
            # sorted_months = sorted(group_months, key=month_lookup.index)
            # converted_months = []
            options = ""
            for month in month_list:
                # converted_months = month.replace('January', 'Yanvar').replace('February','Fevral').replace('March','Mart').replace('April','Aprel').replace('June','Iyun').replace('July','Iyul').replace('August','Avgust').replace('September','Sentyabr').replace('October','Oktyabr').replace('November','Noyabr').replace('December','Dekabr')
                options += f"<option value='{month}'>{month}</option>"
            return HttpResponse(options)
    else:
        return False


class SmsSettings(LoginRequiredMixin, ListView):
    model = Sms
    template_name = 'user/director/sms_settings.html'
    context_object_name = 'send_sms'
    paginate_by = 10
    queryset = Sms.objects.all().order_by('-id')

    def get_queryset(self):
        return Sms.objects.filter(school=self.request.user.school)

    def get(self, request, *args, **kwargs):
        if request.user.role == '2':
            if request.GET.get('q', None):
                try:
                    qs = Sms.objects.filter(text__icontains=request.GET.get('q'), phone__icontains=request.GET.get('q'),
                                            school=request.user.school)
                except:
                    qs = Sms.objects.filter(text__icontains=request.GET.get('q'), school=request.user.school)
                self.queryset = qs
            return super().get(request, *args, **kwargs)
        else:
            return render(request, 'inc/404.html')

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context.update(SMS_PRICE=SMS_PRICE)
        context.update(ADD_PUPIL_SMS_PRICE=ADD_PUPIL_SMS_PRICE)
        context.update(SMS_ADD_STEP=SMS_ADD_STEP)
        context.update(ADD_PUPIL_SMS_ADD_STEP=ADD_PUPIL_SMS_ADD_STEP)

        for sms in context.get('send_sms'):
            if sms.status == PROCESSING:
                r = GetStatusSms(id=sms.sms_id).get()
                if r == SUCCESS:
                    sms.status = SUCCESS
                elif r == FAILED:
                    sms.status = FAILED
                else:
                    sms.status = PROCESSING
                sms.save()
        return context


class BuySimpleSms(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        try:
            if request.user.role == '2':
                school = get_object_or_404(School, id=request.user.school.id)
                if request.user == school.director:
                    count = int(request.POST.get('count'))
                    if count >= SMS_ADD_STEP:
                        money = count * SMS_PRICE
                        if school.money >= money:
                            school.money -= money
                            school.sms_count += count
                            school.save()

                            BuySms.objects.create(school=school, money=money, sms_count=count)
                            return HttpResponse(True)
                        else:
                            return HttpResponse(False)
                    else:
                        return HttpResponse(False)
                else:
                    return HttpResponse(False)
            else:
                return HttpResponse(False)
        except:
            return HttpResponse(False)


class BuyAddPupilSms(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        try:
            if request.user.role == '2':
                school = get_object_or_404(School, id=request.user.school.id)
                if request.user == school.director:
                    count = int(request.POST.get('count'))
                    if count >= ADD_PUPIL_SMS_ADD_STEP:
                        money = count * ADD_PUPIL_SMS_PRICE
                        if school.money >= money:
                            school.money -= money
                            school.add_pupil_sms_count += count
                            school.save()

                            BuySms.objects.create(school=school, money=money, sms_count=count)
                            return HttpResponse(True)
                        else:
                            return HttpResponse(False)
                    else:
                        return HttpResponse(False)
                else:
                    return HttpResponse(False)
            else:
                return HttpResponse(False)
        except:
            return HttpResponse(False)


@login_required
def modify_checkbox_send_sms(request):
    if request.user.role == '2':
        if request.method == 'POST':
            school = get_object_or_404(School, id=request.user.school.id)
            addPupil = request.POST.get('addPupil')
            editPupil = request.POST.get('editPupil')
            attendancePupil = request.POST.get('attendancePupil')
            paymentPupil = request.POST.get('paymentPupil')
            ratingPupil = request.POST.get('ratingPupil')
            addWorker = request.POST.get('addWorker')
            editWorker = request.POST.get('editWorker')

            if addPupil == 'true':
                school.send_sms_add_pupil = True
            else:
                school.send_sms_add_pupil = False

            if editPupil == 'true':
                school.send_sms_edit_pupil = True
            else:
                school.send_sms_edit_pupil = False

            if attendancePupil == 'true':
                school.send_sms_attendance = True
            else:
                school.send_sms_attendance = False

            if paymentPupil == 'true':
                school.send_sms_payment = True
            else:
                school.send_sms_payment = False

            if ratingPupil == 'true':
                school.send_sms_rating = True
            else:
                school.send_sms_rating = False

            if addWorker == 'true':
                school.send_sms_add_worker = True
            else:
                school.send_sms_add_worker = False

            if editWorker == 'true':
                school.send_sms_edit_worker = True
            else:
                school.send_sms_edit_worker = False

            school.save()
            return HttpResponse(True)
        else:
            return HttpResponse(False)
    else:
        return render(request, 'inc/404.html')


@login_required
def payment_payme(request):
    from clickuz import ClickUz
    from click.models import Order

    order = Order.objects.create(amount=1000, user=request.user)
    url = ClickUz.generate_url(order_id=order.id, amount=order.amount,
                               return_url='http://127.0.0.1:8000/user/payment-payme/')
    print(url)
    if request.POST:
        print('POST')
        print(request.POST)
    if request.GET:
        print("GET")
        print(request.GET)
    return HttpResponse(url)


@login_required
def personal_exam_doc_generate(request, id):
    user = get_object_or_404(User, id=id)

    context = {}

    now = datetime.datetime.now()
    passport_issued_time = dt.strptime(str(user.passport_issued_time), "%Y-%m-%d").strftime("%d.%m.%Y")
    medical_issued_date = dt.strptime(str(user.medical_issued_date), "%Y-%m-%d").strftime("%d.%m.%Y")
    birthday = dt.strptime(str(user.birthday), "%Y-%m-%d").strftime("%d.%m.%Y")

    if user.group.category == 'B':
        if user.school.is_capital:
            doc = DocxTemplate(
                "templates/docs/capital/personal_exam_b.docx")
            # doc = DocxTemplate("H:\django_projects\onless\onless\media\docs\capital\personal_exam_b.docx")
            context.update(
                pupil_name=user.name,
                year=now.year,
                birthday=birthday,
                traffic_rules_teacher=user.group.teacher,
                car_structure_teacher=user.group.car_structure_teacher,
                group_number=user.group.number,
                group_category=user.group.category,
                instructors=user.group.instructors,
                birth_place=user.place_of_birth,
                address=f"{user.region.title.split(' ', 1)[0] if user.region else ''} {user.district.title if user.district else ''} {user.residence_address if user.residence_address else ''}",
                pass_seriya=user.pasport,
                passport_issued_time=passport_issued_time,
                passport_issued_organization=user.passport_issued_organization,
                medical_series=user.medical_series,
                medical_issued_organization=user.medical_issued_organization,
                school=user.school,
                certificate_series=user.certificate_series,
                certificate_number=user.certificate_number,
                medical_issued_date=medical_issued_date
            )

        else:
            doc = DocxTemplate("templates/docs/personal_exam_b.docx")
            # doc = DocxTemplate("H:\django_projects\onless\onless\media\docs\personal_exam_b.docx")
        # doc = DocxTemplate("static/docs/personal_exam_b.docx")

    elif user.group.category == 'BC':
        doc = DocxTemplate("templates/docs/personal_exam_bc.docx")
        # doc = DocxTemplate("static/docs/personal_exam_bc.docx")
    elif user.group.category == 'C':
        doc = DocxTemplate("templates/docs/personal_exam_c.docx")
        # doc = DocxTemplate("static/docs/personal_exam_c.docx")
    elif user.group.category == 'D':
        doc = DocxTemplate("templates/personal_exam_a.docx")
        # doc = DocxTemplate("static/docs/personal_exam_d.docx")
    elif user.group.category == 'E':
        doc = DocxTemplate("templates/docs/personal_exam_e.docx")
        # doc = DocxTemplate("static/docs/personal_exam_e.docx")
    elif user.group.category == 'D':
        doc = DocxTemplate("templates/docs/personal_exam_d.docx")
        # doc = DocxTemplate("static/docs/personal_exam_d.docx")
    else:
        doc = DocxTemplate("templates/docs/personal_exam.docx")
        # doc = DocxTemplate("static/docs/personal_exam.docx")

    context.update(
        name=user.name,
        year=now.year,
        birthday=birthday,
        teacher=f"{user.group.teacher.name.upper() if user.group.teacher else ''}",
        group=f"{user.group.category}-{user.group.number} {user.group.year} o'quv guruhi",
        group_number=user.group.number,
        group_category=user.group.category,
        place_of_birth=f"{user.place_of_birth.upper() if user.place_of_birth else ''}",
        residence_address=f"{user.region.title.split(' ', 1)[0] if user.region else ''} {user.district.title.upper() if user.district else ''} {user.residence_address.upper() if user.residence_address else ''}",
        passport_series_and_date=f"{f'{user.pasport}  {passport_issued_time}' if user.pasport else ''}",
        passport_org=f"{f'{user.passport_issued_organization.upper()}' if user.passport_issued_organization else ''}",
        medical_sery_and_date=f"{f'№{user.medical_series} {medical_issued_date}' if user.medical_series else ''}",
        medical_organization=f"{user.medical_issued_organization.upper() if user.medical_issued_organization else ''}",
        school=user.school,
        certificate_sery_and_number=f"{f'{user.certificate_series.upper()} {user.certificate_number}' if user.certificate_series and user.certificate_number else ''}",

    )

    doc.render(context)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    filename = "{0} - imtihon varaqa.docx".format(user.name.replace('’', "'"))
    content = "attachment; filename=%s" % (filename)
    response['Content-Disposition'] = content
    doc.save(response)
    return response


@login_required
def again_send_sms(request, pupil_id):
    try:
        pupil = get_object_or_404(User, id=pupil_id)
        school = get_object_or_404(School, id=pupil.school.id)

        msg = f"Hurmatli {pupil.name}! Siz {pupil.group.category}-{pupil.group.number} guruhiga o'qishga qabul qilindingiz. Videodarslarni ko'rish va imtihon testlariga tayyorlanish uchun http://onless.uz/kirish manziliga kiring. %0aLogin: {pupil.username}%0aParol: {pupil.turbo}%0aQo'shimcha ma'lumot uchun:%0a{pupil.school.phone}"
        msg = msg.replace(" ", "+")
        url = f"https://developer.apix.uz/index.php?app=ws&u=jj39k&h=cb547db5ce188f49c1e1790c25ca6184&op=pv&to=998{pupil.phone}&msg={msg}"
        response = requests.get(url)
        return HttpResponse(f"{pupil.group.category}-{pupil.group.number} {pupil.group.year}: {pupil.name}")
    except:
        return HttpResponse('Error send again sms. PupilId: ' + pupil_id)


def error_403(request):
    return render(request, 'inc/404.html')


def handler404(request, exception):
    return render(request, 'inc/404.html', status=404)
